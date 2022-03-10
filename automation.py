import openpyxl

#criar uma planilha (book)

book=openpyxl.Workbook()
#visualizar paginas exitentes na planilha
print(book.sheetnames)
#como rodar a pagina
book.create_sheet('frutas')
#como delecionar uma pagina
frutas_page = book['frutas']
frutas_page.append(['Frutas','Quatidade','Valor'])
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Uva','5','R$15,90'])
frutas_page.append(['Maçã','5','R$30,90'])
frutas_page.append(['Goiaba','5','R$50,50'])
#salvar a planilha
book.save('Planilha de Compras.xlsx')
